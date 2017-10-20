import os
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font
import tkinter as tk
from tkinter import *

#district list
DList = ['BANKURA','BIRBHUM','BURDWAN','COOCHBEHAR','DAKSHIN DINAJPUR','DARJEELING','HOOGHLY','HOWRAH','JALPAIGURI','MALDA','MURSHIDABAD','NADIA','NORTH 24 PARGANAS','PASCHIM MEDINIPUR','PURBA MEDINIPUR','PURULIA','SOUTH 24 PARGANAS','UTTAR DINAJPUR']
CropList = ['AUS','AMAN','BORO','WHEAT','MAIZE','JUTE','MUSUR','MASKALAI','KHESARI','GRAM','MUSTARD','TIL','POTATO','SUGARCANE']
years = ['2013-14','2014-15','2015-16']
ParamList = ['Area', 'Production', 'Yield']
allCols = ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z','AA','AB','AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM','AN','AO','AP','AQ','AR','AS','AT','AU','AV','AW','AX','AY','AZ']

def writer (shno,dst,crp,prm,blk,val):
    ws[shno].cell(row=r,column=1).value=dst
    ws[shno].cell(row=r,column=2).value=crp
    ws[shno].cell(row=r,column=3).value=prm
    ws[shno].cell(row=r,column=4).value=blk
    ws[shno].cell(row=r,column=5).value=val

def getList():
    lis = ent.get()
    global DList,year
    if lis != '':
        DList = []
        for wrd in (lis).split(','):
            DList.append(wrd)
    yr = enty.get()
    if yr != '':
        year=[]
        for wrd in (yr).split(','):
            year.append(wrd)
    else:
        tk.messagebox.showerror(title="Error",message="No year entered.\nPlease enter years.")
    root.destroy()
root = tk.Tk()
root.wm_title("Comparison")
lab = Label(root, width=15, text="District names", anchor='w')
ent = Entry(root, width=30)
laby = Label(root, width=15, text="Years", anchor='w')
enty = Entry(root, width=30)
inst = Label(root, justify=LEFT, fg='#656565', text="Enter district names and years separated by commas.\nEnter nothing and press okay to convert\nall districts.\nSeparate year with \'-\', for example \'2014-15\'")
but = Button(root, width=8, text="Okay", command=getList)
lab.grid(row=1,column=1)
laby.grid(row=2,column=1)
ent.grid(row=1,column=2,columnspan=2)
enty.grid(row=2,column=2,columnspan=2)
inst.grid(row=3,column=1,columnspan=2)
but.grid(row=3,column=3)
root.mainloop()

for dist in DList:

    print (dist)

    if not os.path.exists("OUTPUT"):
        os.makedirs("OUTPUT")
        
    wb = Workbook()
    ws = []
    ws.append(wb.active)
    ws[0].title = years[1]+" over "+years[0]
    ws.append(wb.create_sheet(years[2]+" over "+years[1]))
    
    r=1

    writer (0,"District","Crop","Parameter","Block","Percentage Variation")
    for row in ws[0].iter_rows(min_row=1, max_row=1, max_col=5):
       for cell in row:
           cell.font= Font(bold=True)

    writer (1,"District","Crop","Parameter","Block","Percentage Variation")
    for row in ws[1].iter_rows(min_row=1, max_row=1, max_col=5):
       for cell in row:
           cell.font= Font(bold=True)

    wkb=[]
    wks=[]
    
    for year in years:
        
        cwb = load_workbook(filename="INPUT\\"+year+"\\"+dist.replace(" ","_")+"_Crop_"+year+".xlsx", read_only=True)
        wkb.append(cwb)
        wks.append(cwb['18.1'])

    for crop in CropList:

        print (" "+crop)

        for k in range (0,3):
            for i in range(0,52):
                col = allCols[i]
                if wks[0][col+str(3)].value!= None:
                    if crop in wks[0][col+str(3)].value.upper():
                        for j in range (1,32):
                            nm=None
                            for m in range (0,3):
                                if wks[m]['B'+str(5+j)].value != None:
                                    nm=wks[m]['B'+str(5+j)].value
                            if nm != None:
                                r+=1
                                for l in range (0,2):
                                    if (wks[0+l][allCols[i+k]+str(5+j)].value=="ERROR")|(wks[1+l][allCols[i+k]+str(5+j)].value=="ERROR"):
                                        writer (l,dist,crop,ParamList[k],nm,"ERROR IN COMPARISON")
                                    
                                    elif wks[0+l][allCols[i+k]+str(5+j)].value==None:
                                        if wks[1+l][allCols[i+k]+str(5+j)].value==None:
                                            writer (l,dist,crop,ParamList[k],nm,0)
                                        elif float(wks[1+l][allCols[i+k]+str(5+j)].value)==0:
                                            writer (l,dist,crop,ParamList[k],nm,0)
                                        else:
                                            writer (l,dist,crop,ParamList[k],nm,"Change from 0")
                                    elif wks[1+l][allCols[i+k]+str(5+j)].value==None:
                                        if float(wks[0+l][allCols[i+k]+str(5+j)].value)==0:
                                            writer (l,dist,crop,ParamList[k],nm,0)
                                        else:
                                            writer (l,dist,crop,ParamList[k],nm,"Change to 0")
                                    elif float(wks[0+l][allCols[i+k]+str(5+j)].value)==0:
                                        if wks[1+l][allCols[i+k]+str(5+j)].value==0:
                                            writer (l,dist,crop,ParamList[k],nm,0)
                                        else:
                                            writer (l,dist,crop,ParamList[k],nm,"Change from 0")
                                    elif float(wks[1+l][allCols[i+k]+str(5+j)].value)==0:
                                        writer (l,dist,crop,ParamList[k],nm,"Change to 0")
                                    elif ((float(wks[0+l][allCols[i+k]+str(5+j)].value)!= 0)&(float(wks[1+l][allCols[i+k]+str(5+j)].value)!=0))&((wks[0+l][allCols[i+k]+str(5+j)].value!=None)&(wks[1+l][allCols[i+k]+str(5+j)].value!=None)):
                                        delta = (float(wks[1+l][allCols[i+k]+str(5+j)].value)-float(wks[0+l][allCols[i+k]+str(5+j)].value))/float(wks[0+l][allCols[i+k]+str(5+j)].value)*100
                                        writer (l,dist,crop,ParamList[k],nm,delta)

    wb.save("OUTPUT\\"+dist+".xlsx")